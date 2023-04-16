#!/usr/bin/env python3

# multiplicationTable.py - Takes a number N from the command-line and creates
# an NxN multiplication table in an Excel spreadsheet. Saves the created table
# file in the user's home directory.
# Row 1 and Column A are used for labels and are in bold.


import sys
import openpyxl
from openpyxl.styles import Font
import os


def create_table(num):
    print('\nCreating a {}x{} table...'.format(num, num))
    wb = openpyxl.Workbook()    # Create a blank workbook.
    sheet = wb['Sheet']

    # Loop through the rows and insert the table labels into Column A and Row 1
    for i in range(2, (num + 2)):
        boldFont = Font(bold=True)
        sheet.cell(row=i, column=1,).font = boldFont    # Set bold font.
        sheet.cell(row=1, column=i,).font = boldFont
        sheet.cell(row=i, column=1,).value = i - 1      # Insert label into Column A.
        sheet.cell(row=1, column=i,).value = i - 1      # Insert label into Row 1.

    # Loop through the rows and columns and insert the value of Column x Row.
    for i in range(2, (num + 2)):
        for j in range(2, (num + 2)):
            multiplyValue = int(sheet.cell(row=j, column=1).value) * int(sheet.cell(row=1, column=i).value)
            sheet.cell(row=j, column=i).value = multiplyValue

    # Get user home directory to save file.
    home = os.path.expanduser('~')
    filename = '{}x{}-table.xlsx'.format(num, num)
    file = os.path.join(home, filename)
    print('Saving file in {} as {}...'.format(home, filename))
    wb.save(file)


def main():
    if len(sys.argv) != 2:
        print('\nUsage: multiplicationTable.py <integer>')
        sys.exit(1)
    
    try:
        number = int(sys.argv[1])
        create_table(number)
    except ValueError:
        print('\nPlease use a whole number (integer)...')
        sys.exit(1)


if __name__ == "__main__":
    main()

