'''
Chapter 16 Project: Excel-to-CSV Converter

excel_to_csv.py - Takes a directory path as a command-line argument and searches
the given directory for any '.xlsx' (Excel) files and converts each file/sheet
into a CSV file. Example: spreadsheet.xlsx would be converted into the files
spreadsheet-<sheet title>.csv.
'''


import sys
import os
from pathlib import Path
import csv
import openpyxl


def read_excel_spreadsheets(directory):
    """Searches for Excel files in the given directory and reads the data from them.

    Args:
        directory (str): Path to a directory to search for Excel files in.
    """
    print(f'\nSearching {directory} for .xlsx files...')

    # Read in the spreadsheet data for each sheet that contains data in the file
    # as a list of lists and create a new filename for the CSV file.
    p = Path(directory)
    for spreadsheet_file in os.listdir(directory):
        # Skip non-xlsx files, load the workbook object.
        if (p / spreadsheet_file).suffix != '.xlsx':
            continue
        else:
            file = p / spreadsheet_file
            wb = openpyxl.load_workbook(file)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                # Create the name for the new file of <spreadsheet file>-<sheet title>.csv.
                new_filename = str(str(p / file.stem) + '-' + sheet.title + '.csv')
                print(f'\nWriting {new_filename}')

                # Loop through every row in the sheet.
                for row_num in range(1, sheet.max_row + 1):
                    row_data = []
                    # Loop through each cell in the row.
                    for col_num in range(1, sheet.max_column + 1):
                        row_data.append(sheet.cell(row=row_num, column=col_num).value)

                    write_csv_file(new_filename, row_data)


def write_csv_file(filename_path, list_data):
    """Writes the given data to a CSV file with the given filename and path.

    Args:
        filename_path (str): Absolute path and filename for the new CSV file.
        list_data (list): List of data to add to the CSV file.
    """
    # Write the data into the new CSV file.
    with open(filename_path, 'a',encoding='UTF-8', newline='') as file:
        writer_obj = csv.writer(file)
        writer_obj.writerow(list_data)


def main():
    """Main function to run the program.
    """
    # Ensure there is only one command line argument given and that
    # the given directory exists.
    if len(sys.argv) != 2:
        print('\nUsage: excel_to_csv.py </path/to/excel/files>')
        sys.exit(1)

    p = str(sys.argv[1])
    if not Path(p).exists():
        print(f'\nPath {p} does not exist!')
        sys.exit(1)
    elif not Path(p).is_dir():
        print(f'\nPath {p} does not point to a directory!')

    read_excel_spreadsheets(p)


if __name__ == "__main__":
    main()
