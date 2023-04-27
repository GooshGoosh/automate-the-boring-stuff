#!/usr/bin/env python3

# spreadsheetToText.py - Takes the path to a spreadsheet file as a command-line argument
# and parses each column of the spreadsheet into separate text files. The created text
# files are placed into the same directory as the spreadsheet file in the format
# <spreadsheet name>-column-<column letter>.txt

# For example, a spreadsheet with columns A, B, and C will produce 3 separate text files
# with each row in the columns being a separate line of text in the respective text files.


import openpyxl
import sys
from pathlib import Path
from openpyxl.utils import get_column_letter


def read_spreadsheet(spreadsheetFile):
    validExtensions = ['.xlsx', '.xlsm', '.xltx', '.xltm'] # List of valid Excel extensions.

    # Set the path of the spreadsheet and ensure it exists.
    spreadsheet = Path(spreadsheetFile)
    if not spreadsheet.exists():
        print('\nFile path {} does not exist!\nExiting...'.format(spreadsheet))
        sys.exit(1)
    elif not spreadsheet.is_file():
        print('\nThe path {} is not a file!\nExiting...'.format(spreadsheet))
        sys.exit(1)
    elif spreadsheet.suffix not in validExtensions:
        print('\nFile {} is not a valid spreadsheet file!\nExiting...'.format(spreadsheet.name))
        sys.exit(1)

    print('\nLoading workbook...')
    wb = openpyxl.load_workbook(spreadsheet)
    sheet = wb.worksheets[0]    # Load the first worksheet in the workbook.
    
    print('Reading workbook...')
    spreadsheetData = {}    # Dictionary for spreadsheet data.
    # Loop through each column in the spreadsheet.
    for i in range(1, sheet.max_column + 1):
        rowLines = []   # List to hold the data of each row in the column.
        letter = get_column_letter(i)

        # Loop through each row in the column and add the row data to rowLines.
        for j in range(1, len(sheet[letter]) + 1):
            rowLines.append(str(sheet.cell(row=j, column=i).value))

        spreadsheetData[letter] = rowLines   # Add the column and row data to the dictionary.

    wb.close()
    return spreadsheetData


def write_text_files(dictData, spreadsheetPath):
    parentPath = Path(spreadsheetPath).parent   # Set the parent path.
    filename = str(Path(spreadsheetPath).stem)  # Get the basename of the spreadsheet.

    # Loop through the dictionary of data.
    for k in dictData.keys():
        newFilename = filename + '-column-' + k + '.txt'    # Create a new filename.
        with open(parentPath / newFilename, 'w') as file:   # Open a text file using the parent
            for line in dictData[k]:                        # path and new filename.
                file.write(line + '\n')     # Loop through each list in the dictionary
                                            # and write each index to a line in the file.

    print('Text files saved in {}'.format(parentPath))


def main():
    if len(sys.argv) != 2:
        print('\nUsage: spreadsheetToText.py <path/to/spreadsheet>')
        sys.exit(1)

    spreadsheet = sys.argv[1]
    spreadsheetData = read_spreadsheet(spreadsheet)
    
    write_text_files(spreadsheetData, spreadsheet)

if __name__ == "__main__":
    main()

