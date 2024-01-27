'''
Chapter 13 Project: Spreadsheet Cell Inverter

spreadsheet_cell_inverter.py - Inverts the row and column of the cells in the
spreadsheet.
For example, the value of the cell at row 5, column 3 will now be at row 3, column 5.
This is done for all cells.
'''


import sys
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter


def read_spreadsheet(spreadsheet_file):
    """Reads data from a given spreadsheet file.

    Args:
        spreadsheet_file (str): Path to a spreadsheet file.

    Returns:
        dict: Dictionary of spreadsheet data from the given spreadsheet.
    """
    valid_ext = ['.xlsx', '.xlsm', '.xltx', '.xltm']  # List of valid Excel extensions.

    # Set the path of the spreadsheet and ensure it exists.
    spreadsheet = Path(spreadsheet_file)

    if not spreadsheet.exists():
        print(f'\nFile path {spreadsheet} does not exist!\nExiting...')
        sys.exit(1)
    elif not spreadsheet.is_file():
        print(f'\nFile path {spreadsheet} does not exist!\nExiting...')
        sys.exit(1)
    elif spreadsheet.suffix not in valid_ext:
        print(f'\nFile {spreadsheet.name} is not a valid spreadsheet file!\nExiting...')
        sys.exit(1)

    print('\nLoading workbook...')
    wb = openpyxl.load_workbook(spreadsheet)
    sheet = wb.worksheets[0]    # Load the first worksheet in the workbook

    print('Reading workbook...')
    spreadsheet_data = {}        # Dictionary for spreadsheet data.
    max_col = sheet.max_column   # Max column that contains data.

    # Loop through each column in the spreadsheet.
    for i in range(1, max_col + 1):
        col_lines = []   # List to hold the value of each row in the current column.
        letter = get_column_letter(i)
        max_row = len(sheet[letter])

        # Loop through each row in the column and add the data into col_lines.
        for j in range(1, max_row + 1):
            col_lines.append(str(sheet.cell(row=j, column=i).value))

        spreadsheet_data[i] = col_lines   # Add the column and row data to the dictionary.

    wb.close()
    return spreadsheet_data


def invert_spreadsheet(dict_data, spreadsheet_path):
    """Inverts the data from the given dictionary of spreadsheet data and writes
    it to a new spreadsheet file.

    Args:
        dict_data (dict): Dictionary data from a spreadsheet file.
        spreadsheet_path (str): Path to the spreadsheet file being inverted.
    """
    parent_path = Path(spreadsheet_path).parent     # Set the parent path.
    filename = str(Path(spreadsheet_path).stem)     # Get the basename of the spreadsheet.

    print('Inverting spreadsheet...')
    wb = openpyxl.Workbook()    # Create a new workbook.
    sheet = wb.worksheets[0]    # Load the first worksheet in the workbook.

    # Loop through the dictionary of data and add it to the spreadsheet.
    for k, v in dict_data.items():
        # Loop through the list at each dictionary value.
        for i in range(1, len(v) + 1):
            sheet.cell(row=k, column=i).value = v[i - 1]

    new_filename = filename + '-inverted.xlsx'   # Create a new filename.
    wb.save(parent_path / new_filename)

    print(f'File saved in {parent_path} as {new_filename}')


def main():
    """Main function to run the program.
    """
    if len(sys.argv) != 2:
        print('\nUsage: spreadsheetCellInverter.py <path/to/spreadsheet')
        sys.exit(1)

    spreadsheet = sys.argv[1]
    spreadsheet_data = read_spreadsheet(spreadsheet)

    invert_spreadsheet(spreadsheet_data, spreadsheet)


if __name__ == "__main__":
    main()
