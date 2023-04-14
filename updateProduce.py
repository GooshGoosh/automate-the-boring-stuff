#!/usr/bin/env python3

# updateProduce.py - Corrects costs in produce sales spreadsheet.

import openpyxl
from pathlib import Path


def update_produce(spreadsheet):
    print('\nOpening workbook...')
    wb = openpyxl.load_workbook(spreadsheet)
    sheet = wb['Sheet']

    # The produce types and their updated prices.
    priceUpdates = {'Garlic': 3.07,
                    'Celery': 1.19,
                    'Lemon': 1.27}

    # Loop through the rows and update their prices.
    print('Updating prices...')
    for rowNum in range(2, sheet.max_row):  # Skip the first row
        produceName = sheet.cell(row=rowNum, column=1).value
        if produceName in priceUpdates:
            sheet.cell(row=rowNum, column=2).value = priceUpdates[produceName]

    # Create new file name to save the updated file.
    fileParent = Path(spreadsheet).parent   # Directory that the file is in.
    fileStem = Path(spreadsheet).stem       # File name without extension.
    newFile = Path(fileParent).joinpath(fileStem + '-updated.xlsx')

    wb.save(newFile)
    print('Done.')


def main():
    file = Path('/home/alexg/scripts/automate-online-materials/produceSales.xlsx')
    update_produce(file)


if __name__ == "__main__":
    main()

