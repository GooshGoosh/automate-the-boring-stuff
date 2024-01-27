'''
Chapter 13 Project: Updating a Spreadsheet

update_produce.py - Corrects costs in produce sales spreadsheet.
'''


from pathlib import Path
import openpyxl


def update_produce(spreadsheet):
    """Update the data in the give spreadsheet and writes the changes to a new
    spreadsheet file.

    Args:
        spreadsheet (str): Path to a spreadsheet file.
    """
    print('\nOpening workbook...')
    wb = openpyxl.load_workbook(spreadsheet)
    sheet = wb['Sheet']

    # The produce types and their updated prices.
    price_updates = {'Garlic': 3.07,
                    'Celery': 1.19,
                    'Lemon': 1.27}

    # Loop through the rows and update their prices.
    print('Updating prices...')
    for row_num in range(2, sheet.max_row):  # Skip the first row
        produce_name = sheet.cell(row=row_num, column=1).value
        if produce_name in price_updates:
            sheet.cell(row=row_num, column=2).value = price_updates[produce_name]

    # Create new file name to save the updated file.
    file_parent = Path(spreadsheet).parent   # Directory that the file is in.
    file_stem = Path(spreadsheet).stem       # File name without extension.
    new_file = Path(file_parent).joinpath(file_stem + '-updated.xlsx')

    wb.save(new_file)
    print('Done.')


def main():
    """The main function to run the program.
    """
    file = Path('automate-online-materials/produceSales.xlsx')
    update_produce(file)


if __name__ == "__main__":
    main()
