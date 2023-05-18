#!/usr/bin/env python3

# spreadsheetCellInverter.py - Inverts the row and column of the cells in the spreadsheet.
# For example, the value of the cell at row 5, column 3 will now be at row 3, column 5.
# This is done for all cells.


import openpyxl
import sys
from pathlib import Path
from openpyxl.utils import get_column_letter


def read_spreadsheet(spreadsheetFile):
    validExtensions = ['.xlsx', '.xlsm', '.xltx', '.xltm']  # List of valid Excel extensions.

    # Set the path of the spreadsheet and ensure it exists.
    spreadsheet = Path(spreadsheetFile)

    if not spreadsheet.exists():
        print('\nFile path {} does not exist!\nExiting...'.format(spreadsheet))
        sys.exit(1)
    elif not spreadsheet.is_file():
        print('\nFile path {} does not exist!\nExiting...'.format(spreadsheet))
        sys.exit(1)
    elif spreadsheet.suffix not in validExtensions:
        print('\nFile {} is not a valid spreadsheet file!\nExiting...'.format(spreadsheet.name))
        sys.exit(1)

    print('\nLoading workbook...')
    wb = openpyxl.load_workbook(spreadsheet)
    sheet = wb.worksheets[0]    # Load the first worksheet in the workbook

    print('Reading workbook...')
    spreadsheetData = {}        # Dictionary for spreadsheet data.
    maxCol = sheet.max_column   # Max column that contains data.

    # Loop through each column in the spreadsheet.
    for i in range(1, maxCol + 1):
        colLines = []   # List to hold the value of each row in the current column.
        letter = get_column_letter(i)   # Get the letter of the current column.
        maxRow = len(sheet[letter])     # Get the number of rows that contain data.

        # Loop through each row in the column and add the data into colLines.
        for j in range(1, maxRow + 1):
            colLines.append(str(sheet.cell(row=j, column=i).value))

        spreadsheetData[i] = colLines   # Add the column and row data to the dictionary.

    wb.close()
    return spreadsheetData


def invert_spreadsheet(dictData, spreadsheetPath):
    parentPath = Path(spreadsheetPath).parent   # Set the parent path.
    filename = str(Path(spreadsheetPath).stem)  # Get the basename of the spreadsheet.

    print('Inverting spreadsheet...')
    wb = openpyxl.Workbook()    # Create a new workbook.
    sheet = wb.worksheets[0]    # Load the first worksheet in the workbook.

    # Loop through the dictionary of data and add it to the spreadsheet.
    for k, v in dictData.items():
        # Loop through the list at each dictionary value.
        for i in range(1, len(v) + 1):
            sheet.cell(row=k, column=i).value = v[i - 1]

    newFilename = filename + '-inverted.xlsx'   # Create a new filename.
    wb.save(parentPath / newFilename)

    print('File saved in {} as {}'.format(parentPath, newFilename))


def main():
    if len(sys.argv) != 2:
        print('\nUsage: spreadsheetCellInverter.py <path/to/spreadsheet')
        sys.exit(1)

    spreadsheet = sys.argv[1]
    spreadsheetData = read_spreadsheet(spreadsheet)

    invert_spreadsheet(spreadsheetData, spreadsheet)


if __name__ == "__main__":
    main()

