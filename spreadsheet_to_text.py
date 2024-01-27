'''
Chapter 13 Project: Spreadsheet to Text Files

spreadsheet_to_text.py - Takes the path to a spreadsheet file as a command line argument
and parses each column of the spreadsheet into separate text files. The created text
files are placed into the same directory as the spreadsheet file in the format
<spreadsheet name>-column-<column letter>.txt

For example, a spreadsheet with columns A, B, and C will produce 3 separate text files
with each row in the columns being a separate line of text in the respective text files.
'''


import sys
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter


def read_spreadsheet(spreadsheet_file):
    """Reads the data from each sheet in the spreadsheet file.

    Args:
        spreadsheet_file (str): Path to a spreadsheet file.

    Returns:
        dict: Dictionary of spreadsheet data for each sheet.
    """
    valid_ext = ['.xlsx', '.xlsm', '.xltx', '.xltm'] # List of valid Excel extensions.

    # Set the path of the spreadsheet and ensure it exists.
    spreadsheet = Path(spreadsheet_file)
    if not spreadsheet.exists():
        print(f'\nFile path {spreadsheet} does not exist!\nExiting...')
        sys.exit(1)
    elif not spreadsheet.is_file():
        print(f'\nThe path {spreadsheet} is not a file!\nExiting...')
        sys.exit(1)
    elif spreadsheet.suffix not in valid_ext:
        print(f'\nFile {spreadsheet.name} is not a valid spreadsheet file!\nExiting...')
        sys.exit(1)

    print('\nLoading workbook...')
    wb = openpyxl.load_workbook(spreadsheet)
    sheet = wb.worksheets[0]

    print('Reading workbook...')
    spreadsheet_data = {}
    # Loop through each column in the spreadsheet.
    for i in range(1, sheet.max_column + 1):
        row_lines = []
        letter = get_column_letter(i)

        # Loop through each row in the column and add the row data to row_lines.
        for j in range(1, len(sheet[letter]) + 1):
            row_lines.append(str(sheet.cell(row=j, column=i).value))

        spreadsheet_data[letter] = row_lines

    wb.close()
    return spreadsheet_data


def write_text_files(dict_data, spreadsheet_path):
    """Writes the data from the given dictionary to a text file and saves it in
    the path given.

    Args:
        dict_data (dict): Dictionary of spreadsheet data.
        spreadsheet_path (str): Path to a spreadsheet file.
    """
    parent_path = Path(spreadsheet_path).parent   # Set the parent path.
    filename = str(Path(spreadsheet_path).stem)  # Get the basename of the spreadsheet.

    # Loop through the dictionary of data.
    for k in dict_data.keys():
        new_filename = filename + '-column-' + k + '.txt'    # Create a new filename.
        # Open a text file using the parent path and new filename.
        with open(parent_path / new_filename, 'w', encoding='UTF-8') as file:
            for line in dict_data[k]:
                file.write(line + '\n')

    print(f'Text files saved in {parent_path}')


def main():
    """Main function to run the program.
    """
    if len(sys.argv) != 2:
        print('\nUsage: spreadsheetToText.py <path/to/spreadsheet>')
        sys.exit(1)

    spreadsheet = sys.argv[1]
    spreadsheet_data = read_spreadsheet(spreadsheet)

    write_text_files(spreadsheet_data, spreadsheet)


if __name__ == "__main__":
    main()
