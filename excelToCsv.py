#!/usr/bin/env python3

# excelToCsv.py - Takes a directory path as a command-line argument and searches
# the given directory for any '.xlsx' (Excel) files and converts each file/sheet
# into a CSV file. Example: spreadsheet.xlsx would be converted into the files
# spreadsheet-<sheet title>.csv.


import csv
import openpyxl
import sys
import os
from pathlib import Path


def read_excel_spreadsheets(directory):
    # Search the directory for any .xlsx files.
    print('\nSearching {} for .xlsx files...'.format(directory))

    # Read in the spreadsheet data for each sheet that contains data in the file
    # as a list of lists and create a new filename for the CSV file.
    p = Path(directory)
    for spreadsheetFile in os.listdir(directory):
        # Skip non-xlsx files, load the workbook object.
        if (p / spreadsheetFile).suffix != '.xlsx':
            continue
        else:
            file = p / spreadsheetFile
            wb = openpyxl.load_workbook(file)
            for sheetName in wb.sheetnames:
                # Loop through every sheet in the workbook.
                sheet = wb[sheetName]
                # Create the name for the new file of <spreadsheet file>-<sheet title>.csv.
                newFilename = str(str(p / file.stem) + '-' + sheet.title + '.csv')
                print('\nWriting {}'.format(newFilename))

                # Loop through every row in the sheet.
                for rowNum in range(1, sheet.max_row + 1):
                    rowData = []    # Append each cell to this list
                    # Loop through each cell in the row.
                    for colNum in range(1, sheet.max_column + 1):
                        # Append each cell's data to rowData
                        rowData.append(sheet.cell(row=rowNum, column=colNum).value)
                    
                    write_csv_file(newFilename, rowData)


def write_csv_file(filenamePath, listData):
    # Write the data into the new CSV file.
    # Create the csv.writer object for the CSV file.
    with open(filenamePath, 'a', newline='') as file:
        writerObj = csv.writer(file)
        # Write the listData list to the CSV file.
        writerObj.writerow(listData)


def main():
    # Ensure there is only one command-line argument given and that
    # the given directory exists.
    if len(sys.argv) != 2:
        print('\nUsage: excelToCsv.py </path/to/excel/files>')
        sys.exit(1)
    
    p = str(sys.argv[1])
    if not Path(p).exists():
        print('\nPath {} does not exist!'.format(p))
        sys.exit(1)
    elif not Path(p).is_dir():
        print('\nPath {} does not point to a directory!'.format(p))

    read_excel_spreadsheets(p)


if __name__ == "__main__":
    main()

